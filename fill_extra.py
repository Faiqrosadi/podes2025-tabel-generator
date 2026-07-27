#!/usr/bin/env python3
"""
Lengkapi SEMUA bab (1-7) untuk 16 kecamatan:
  - Bab 1,2,3,5: buat hasil/<kec>/Bab N.xlsx dari template, isi data yg tersedia.
  - Bab 4,6,7  : pakai hasil yg sudah diisi (dari fill_bab4/6/7.py).
  - Semua bab  : sel data yang TIDAK ada sumbernya diisi "-" (no data),
                 supaya setiap tabel tetap tampil lengkap.

Data tersedia yang diisi di sini:
  1.1.2  Jarak ke ibukota kec/kab  -> r802ak5, r802bk5   (per desa)
  2.1.1  RW / RT                   -> r304a, r304b        (per desa)
  2.2.5  Perangkat desa            -> Petinggi=r1201ak2==1, Carik=r1201bk2==1,
                                      Perangkat Lainnya=r1202a+b+c+d
Jalankan SETELAH fill_bab4.py, fill_bab6.py, fill_bab7.py.
"""
import os, re, glob, shutil
import pyreadstat, openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.cell.cell import MergedCell

def safe_set(ws, coord, value):
    """Set nilai; lewati bila sel bagian dari merge (read-only)."""
    if not isinstance(ws[coord], MergedCell):
        ws[coord] = value

BASE = os.path.dirname(os.path.abspath(__file__))
try:
    from config import SAV, TEMPLATE as TPL, OUTPUT as OUT
except Exception:
    SAV  = os.path.join(BASE, "data", "33_podes2025-desa_rev_20.sav")
    TPL  = os.path.join(BASE, "template tabel")
    OUT  = os.path.join(BASE, "hasil")

df, _ = pyreadstat.read_sav(SAV)
df["r104"] = df["r104"].astype(str)

def norm(s): return re.sub(r"\s+", " ", str(s).strip()).upper()
def fmt(v):
    try: v = float(v)
    except (TypeError, ValueError): return "-"
    return "-" if (v != v or v == 0) else (int(v) if v == int(v) else v)
def numval(v):
    try:
        v = float(v); return None if v != v else v
    except (TypeError, ValueError): return None

def resolve_formulas(ws):
    """Hitung rumus Excel (SUM/aritmetika) & tulis nilainya. Cache kosong -> None."""
    memo = {}
    def split_ref(ref):
        m = re.match(r"([A-Za-z]+)(\d+)", ref.replace("$", "").strip())
        return column_index_from_string(m.group(1)), int(m.group(2))
    def sum_range(rng):
        total = 0.0
        for part in rng.split(","):
            part = part.replace("$", "").strip()
            if not part: continue
            if ":" in part:
                a, b = part.split(":"); c1, r1 = split_ref(a); c2, r2 = split_ref(b)
                for c in range(min(c1, c2), max(c1, c2) + 1):
                    for r in range(min(r1, r2), max(r1, r2) + 1):
                        total += num(f"{get_column_letter(c)}{r}")
            else:
                total += num(part)
        return total
    def eval_expr(expr):
        expr = expr.replace("$", "")
        expr = re.sub(r"(?i)sum\(([^)]*)\)", lambda m: str(sum_range(m.group(1))), expr)
        expr = re.sub(r"[A-Za-z]+\d+", lambda m: str(num(m.group(0))), expr)
        try:
            return float(eval(expr, {"__builtins__": {}}, {}))
        except Exception:
            return None
    def num(coord):
        coord = coord.replace("$", "")
        if coord in memo: return memo[coord] or 0.0
        memo[coord] = 0.0
        v = ws[coord].value
        if isinstance(v, (int, float)):
            memo[coord] = float(v)
        elif isinstance(v, str) and v.strip().startswith("="):
            r = eval_expr(v.strip()[1:]); memo[coord] = r if r is not None else 0.0
        return memo[coord] or 0.0

    cells = [c.coordinate for row in ws.iter_rows() for c in row
             if isinstance(c.value, str) and c.value.strip().startswith("=")]
    results = {}
    for coord in cells:
        v = ws[coord].value.strip()[1:]
        results[coord] = eval_expr(v)
    for coord, val in results.items():
        if val is None or val == 0:
            out = "-"
        elif val == int(val):
            out = int(val)
        else:
            out = round(val, 2)
        safe_set(ws, coord, out)

def sheet_by_num(wb, tid):
    for s in wb.sheetnames:
        m = re.search(r"(\d+(?:\.\d+)+)", s)
        if m and m.group(1) == tid: return wb[s]
    return None

# ---------- pengisi data spesifik (per desa) ----------
def fill_per_desa(ws, sub, colmap, do_total=True):
    """colmap = {col_letter: fungsi(row)->nilai}. Baris dicocokkan via kode desa di kolom A.
    Bila do_total, baris kecamatan (nama kec / Jumlah / Total) diisi penjumlahan desa."""
    by_code = {str(r.r104): r for r in sub.itertuples()}
    kec = norm(sub.iloc[0].nama_kec)
    sums = {cl: 0.0 for cl in colmap}
    seen = False
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None: continue
        a = str(a).strip()
        if re.fullmatch(r"\d{3}", a):
            row = by_code.get(a)
            if row is None: continue
            seen = True
            for cl, fn in colmap.items():
                val = fn(row)
                safe_set(ws, f"{cl}{r}", fmt(val))
                v = numval(val)
                if v: sums[cl] += v
        elif do_total and seen and (norm(a) == kec or "JUMLAH" in norm(a) or "TOTAL" in norm(a)):
            for cl in colmap:
                safe_set(ws, f"{cl}{r}", fmt(sums[cl]))
            break

def bab1(wb, sub):   # 1.1.2 jarak (tak dijumlah -> total dibiarkan '-')
    ws = sheet_by_num(wb, "1.1.2")
    if ws: fill_per_desa(ws, sub, {
        "D": lambda r: numval(r.r802ak5),
        "E": lambda r: numval(r.r802bk5)}, do_total=False)

def bab2(wb, sub):   # 2.1.1 RW/RT, 2.2.5 perangkat (baris kec = penjumlahan)
    ws = sheet_by_num(wb, "2.1.1")
    if ws: fill_per_desa(ws, sub, {
        "D": lambda r: numval(r.r304a),
        "E": lambda r: numval(r.r304b)})
    ws = sheet_by_num(wb, "2.2.5")
    if ws: fill_per_desa(ws, sub, {
        "D": lambda r: 1 if numval(r.r1201ak2) == 1 else 0,
        "E": lambda r: 1 if numval(r.r1201bk2) == 1 else 0,
        "F": lambda r: sum((numval(getattr(r, v)) or 0)
                           for v in ("r1202a","r1202b","r1202c","r1202d"))})

SPECIFIC = {1: bab1, 2: bab2}

# ---------- pengisi "-" generik untuk sel no-data ----------
def dashfill(ws):
    """Isi '-' pada sel data yang masih kosong (di bawah baris penanda '(n)')."""
    last_r = ws.max_row
    covered = set()   # sel yg ditutup merge (bukan anchor) -> read-only
    for m in ws.merged_cells.ranges:
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                if (r, c) != (m.min_row, m.min_col): covered.add((r, c))
    # kumpulkan semua baris penanda + kolom nilainya
    markers = []   # (row, [value_cols], first_value_col)
    for r in range(1, last_r + 1):
        nums = {}
        for c in range(1, ws.max_column + 1):
            m = re.fullmatch(r"\((\d+)\)", str(ws.cell(r, c).value).strip()) \
                if ws.cell(r, c).value is not None else None
            if m: nums[c] = int(m.group(1))
        if nums and 1 in nums.values():
            vcols = [c for c, n in nums.items() if n >= 2]
            if vcols:
                markers.append((r, vcols, min(vcols)))
    if not markers: return
    marker_rows = [m[0] for m in markers]
    for idx, (mrow, vcols, first_v) in enumerate(markers):
        end = marker_rows[idx + 1] - 1 if idx + 1 < len(markers) else last_r
        for r in range(mrow + 1, end + 1):
            # baris data bila ada label di kolom sebelum kolom nilai
            has_label = any(ws.cell(r, c).value not in (None, "")
                            for c in range(1, first_v))
            if not has_label: continue
            for c in vcols:
                if (r, c) in covered: continue
                if ws.cell(r, c).value is None:
                    ws.cell(r, c).value = "-"

# ---------- proses ----------
def process(folder):
    kec = norm(re.sub(r"^\d+\s*", "", os.path.basename(folder)))
    sub = df[df["nama_kec"].apply(norm) == kec].sort_values("r104")
    if sub.empty:
        print(f"  ! LEWAT {folder}: nama_kec '{kec}' tak ada"); return
    done = []
    for bab in [1, 2, 3, 4, 5, 6, 7]:
        fname = f"Bab {bab}.xlsx"
        src = os.path.join(folder, fname)
        dst_dir = os.path.join(OUT, os.path.basename(folder)); os.makedirs(dst_dir, exist_ok=True)
        dst = os.path.join(dst_dir, fname)
        if bab in (1, 2, 3, 5):          # buat dari template
            if not os.path.exists(src): continue
            shutil.copy2(src, dst)
        if not os.path.exists(dst):      # bab 4/6/7 harus sudah ada
            print(f"    (bab {bab} belum ada hasil, dilewati)"); continue
        wb = openpyxl.load_workbook(dst)
        if bab in SPECIFIC: SPECIFIC[bab](wb, sub)
        for s in wb.sheetnames:
            resolve_formulas(wb[s])   # hitung rumus (SUM/aritmetika) -> nilai
            dashfill(wb[s])           # sisa sel data kosong -> "-"
        wb.save(dst); done.append(str(bab))
    print(f"  OK {os.path.basename(folder):22} ({len(sub)} desa) bab[{','.join(done)}]")

if __name__ == "__main__":
    folders = [f for f in sorted(glob.glob(os.path.join(TPL, "*"))) if os.path.isdir(f)]
    print(f"Memproses {len(folders)} kecamatan...\n")
    for f in folders: process(f)
    print(f"\nSelesai. Output di: {OUT}")
