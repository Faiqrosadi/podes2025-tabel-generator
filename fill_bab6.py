#!/usr/bin/env python3
"""
Isi otomatis Bab 6 (Akomodasi, Transportasi, Komunikasi) PODES 2025 Kab. Jepara.
Lihat MAPPING.md. Semua nilai = jumlah desa (kolom 2025 / kolom B). 0 -> "-".
Hasil -> hasil/<kecamatan>/Bab 6.xlsx (template asli tak diubah).

  6.1.1  Desa dgn akomodasi     -> count(r905hk2>0) Hotel, count(r905ik2>0) Penginapan
  6.2.1  Desa menurut transport -> count(r801a==kode)  (1 darat,2 air,3 darat&air,4 udara)
  6.3.1  Desa dgn pos/ekspedisi -> Kantor Pos r805a ada(≠4), Pos Keliling r805b==1,
                                    Ekspedisi r805c ada(≠4)
"""
import os, re, glob, shutil
import pyreadstat
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
try:
    from config import SAV, TEMPLATE as TPL, OUTPUT as OUT
except Exception:
    SAV  = os.path.join(BASE, "data", "33_podes2025-desa_rev_20.sav")
    TPL  = os.path.join(BASE, "template tabel")
    OUT  = os.path.join(BASE, "hasil")

df, _ = pyreadstat.read_sav(SAV)
df["r104"] = df["r104"].astype(str)

def norm(s): return re.sub(r"\s+", " ", str(s).strip()).upper()
def fmt(v):
    try: v = float(v)
    except (TypeError, ValueError): return "-"
    return "-" if (v != v or v == 0) else (int(v) if v == int(v) else v)
def cnt(cond): return int(cond.sum())
def num(sub, var): return sub[var].fillna(0)
def sheet(wb, tag):
    for s in wb.sheetnames:
        if tag in s.replace(" ", ""): return wb[s]
    return None
def data_start(ws):
    """Baris pertama data = tepat setelah baris penanda berisi '(1)'."""
    for r in range(1, ws.max_row + 1):
        if any(str(ws.cell(r, c).value).strip() == "(1)"
               for c in range(1, ws.max_column + 1)):
            return r + 1
    return 1

def row_labels(ws, col=1):
    out = {}
    start = data_start(ws)
    for r in range(start, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is not None and str(v).strip() and not str(v).strip().startswith("("):
            out[r] = norm(v)
    return out

def fill_611(ws, sub):   # col B: akomodasi
    for r, t in row_labels(ws).items():
        if "HOTEL" in t:          ws[f"B{r}"] = fmt(cnt(num(sub,"r905hk2") > 0))
        elif "PENGINAPAN" in t:   ws[f"B{r}"] = fmt(cnt(num(sub,"r905ik2") > 0))

def fill_621(ws, sub):   # col B: moda lalu lintas antardesa (r801a)
    for r, t in row_labels(ws).items():
        code = None
        if "DARAT DAN AIR" in t: code = 3
        elif "UDARA" in t:       code = 4
        elif "DARAT" in t:       code = 1
        elif "AIR" in t:         code = 2
        if code:                 ws[f"B{r}"] = fmt(cnt(num(sub,"r801a") == code))

def fill_631(ws, sub):   # col B: pos & ekspedisi
    for r, t in row_labels(ws).items():
        if "KANTOR POS" in t:
            ws[f"B{r}"] = fmt(cnt(num(sub,"r805a").isin([1,2,3])))
        elif "POS KELILING" in t:
            ws[f"B{r}"] = fmt(cnt(num(sub,"r805b") == 1))
        elif "EKSPEDISI" in t:
            ws[f"B{r}"] = fmt(cnt(num(sub,"r805c").isin([1,2,3])))

FILLERS = {"6.1": fill_611, "6.2": fill_621, "6.3": fill_631}

def process(folder):
    kec = norm(re.sub(r"^\d+\s*", "", os.path.basename(folder)))
    sub = df[df["nama_kec"].apply(norm) == kec].sort_values("r104")
    if sub.empty:
        print(f"  ! LEWAT {folder}: nama_kec '{kec}' tak ada"); return False
    src = os.path.join(folder, "Bab 6.xlsx")
    if not os.path.exists(src):
        print(f"  ! LEWAT {folder}: tak ada 'Bab 6.xlsx'"); return False
    dst_dir = os.path.join(OUT, os.path.basename(folder)); os.makedirs(dst_dir, exist_ok=True)
    dst = os.path.join(dst_dir, "Bab 6.xlsx"); shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst); done = []
    for tag, fn in FILLERS.items():
        ws = sheet(wb, tag)
        if ws is not None: fn(ws, sub); done.append(tag)
    wb.save(dst)
    print(f"  OK {os.path.basename(folder):22} ({len(sub)} desa) [{','.join(done)}]")
    return True

if __name__ == "__main__":
    folders = [f for f in sorted(glob.glob(os.path.join(TPL, "*"))) if os.path.isdir(f)]
    print(f"Memproses {len(folders)} kecamatan...\n")
    ok = sum(process(f) for f in folders)
    print(f"\nSelesai: {ok}/{len(folders)} kecamatan. Output di: {OUT}")
