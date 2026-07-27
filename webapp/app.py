#!/usr/bin/env python3
"""Dashboard PODES 2025 Kab. Jepara.
Halaman kecamatan MERENDER langsung file Excel hasil (hasil/<kec>/Bab N.xlsx)
apa adanya — judul, header bertingkat, merge, urutan tabel = persis template.
Kartu ringkasan kabupaten dihitung dari .sav.
Jalankan: PORT=8090 python3 webapp/app.py
"""
import os, re, sys, io, json, glob, shutil, subprocess, zipfile
import pyreadstat
import openpyxl
from openpyxl.utils import get_column_letter
from markupsafe import Markup, escape
from flask import Flask, render_template_string, abort, request, redirect, send_file

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, BASE)
try:
    from config import TEMPLATE as MASTER_TPL, OUTPUT as HASIL
except Exception:
    MASTER_TPL = os.path.join(BASE, "template tabel")
    HASIL = os.path.join(BASE, "hasil")

WORK   = os.path.join(BASE, "workspace"); os.makedirs(WORK, exist_ok=True)
ACTIVE = os.path.join(WORK, "active.json")
GEN_TPL = os.path.join(WORK, "template_gen")     # template hasil-generate dari .sav
FILL_STEPS = ["fill_bab7.py", "fill_bab6.py", "fill_bab4.py", "fill_extra.py"]

def norm(s): return re.sub(r"\s+", " ", str(s).strip()).upper()
def slug(s): return re.sub(r"[^a-z0-9]+", "-", str(s).lower()).strip("-")

# ---- state project (kosong sampai user upload .sav) ----
STATE = {"df": None, "sav": None, "kab": "", "kecs": [], "slug2kec": {}, "kec2folder": {}}

def _kec2folder():
    m = {}
    if os.path.isdir(HASIL):
        for f in os.listdir(HASIL):
            if os.path.isdir(os.path.join(HASIL, f)):
                m[norm(re.sub(r"^\d+\s*", "", f))] = f
    return m

def load_state(sav):
    df, _ = pyreadstat.read_sav(sav)
    df["r104"] = df["r104"].astype(str)
    kecs = [norm(k) for k in df.sort_values("r103").drop_duplicates("r103")["nama_kec"]]
    STATE.update(df=df, sav=sav, kab=str(df["nama_kab"].iloc[0]).title(),
                 kecs=kecs, slug2kec={slug(k): k for k in kecs}, kec2folder=_kec2folder())

def clear_state():
    for p in [HASIL, GEN_TPL, os.path.join(WORK, "data.sav"), ACTIVE,
              os.path.join(BASE, "template tabel (terisi)"),
              os.path.join(BASE, "PODES_tabel_terisi.zip")]:
        if os.path.isdir(p): shutil.rmtree(p, ignore_errors=True)
        elif os.path.exists(p): os.remove(p)
    STATE.update(df=None, sav=None, kab="", kecs=[], slug2kec={}, kec2folder={})

def _run(cmd, env):
    r = subprocess.run(cmd, cwd=BASE, env=env, capture_output=True, text=True)
    if r.returncode != 0:
        raise RuntimeError(f"{' '.join(os.path.basename(c) for c in cmd)}:\n{r.stderr[-1500:]}")

def process_sav(sav):
    """Dari 1 .sav: generate template bersih -> isi pipeline -> muat state."""
    env = os.environ.copy()
    env.update(PODES_SAV=sav, PODES_TEMPLATE=GEN_TPL, PODES_OUTPUT=HASIL)
    if os.path.isdir(HASIL): shutil.rmtree(HASIL, ignore_errors=True)
    if os.path.isdir(GEN_TPL): shutil.rmtree(GEN_TPL, ignore_errors=True)
    _run([sys.executable, os.path.join(BASE, "generate.py"), sav, MASTER_TPL, GEN_TPL], env)
    for s in FILL_STEPS:
        _run([sys.executable, os.path.join(BASE, s)], env)
    with open(ACTIVE, "w") as f: json.dump({"sav": sav}, f)
    load_state(sav)

def _bootstrap():
    if os.path.exists(ACTIVE):
        try:
            d = json.load(open(ACTIVE))
            if d.get("sav") and os.path.exists(d["sav"]) and os.path.isdir(HASIL):
                load_state(d["sav"]); return
        except Exception: pass
_bootstrap()

def sub_of(kec):
    df = STATE["df"]; return df[df["nama_kec"].apply(norm) == kec].sort_values("r104")
def tot(var): return int(STATE["df"][var].fillna(0).sum())

# ---- daftar bab (semua tabel dirender apa adanya dari file hasil) ----
BABS = [
    ("bab1", "Bab 1", "Bab 1 · Letak Geografis", "Bab 1.xlsx"),
    ("bab2", "Bab 2", "Bab 2 · Pemerintahan", "Bab 2.xlsx"),
    ("bab3", "Bab 3", "Bab 3 · Kependudukan", "Bab 3.xlsx"),
    ("bab4", "Bab 4", "Bab 4 · Sosial", "Bab 4.xlsx"),
    ("bab5", "Bab 5", "Bab 5 · Pertanian", "Bab 5.xlsx"),
    ("bab6", "Bab 6", "Bab 6 · Akomodasi, Transportasi & Komunikasi", "Bab 6.xlsx"),
    ("bab7", "Bab 7", "Bab 7 · Ekonomi", "Bab 7.xlsx"),
]

# ---------- render satu sheet Excel -> HTML (setia template) ----------
_CACHE = {}
def load_wb(path):
    mt = os.path.getmtime(path)
    key = (path, mt)
    if key not in _CACHE:
        _CACHE[key] = openpyxl.load_workbook(path, data_only=True)
    return _CACHE[key]

def cellval(v):
    if v is None: return ""
    if isinstance(v, float):
        return str(int(v)) if v == int(v) else f"{v:g}"
    return str(v)

def bounds(ws):
    """Batas grid = sel yg BENAR-BENAR berisi nilai (bukan overhang merge kosong)."""
    last_r = last_c = 0
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and str(cell.value).strip() != "":
                last_r = max(last_r, cell.row); last_c = max(last_c, cell.column)
    return last_r, last_c

def biling(txt):
    """Sel dwibahasa: baris terakhir = Inggris (italic), sisanya Indonesia (tegak)."""
    lines = [l.strip() for l in txt.split("\n") if l.strip()]
    if not lines: return Markup("")
    if len(lines) == 1: return escape(lines[0])
    idp = escape(" ".join(lines[:-1]))
    return Markup(f"{idp}<br><span class=en>{escape(lines[-1])}</span>")

def grid_start_row(ws, last_r):
    """Baris awal grid tabel = baris pertama (r>=2) yg kolom A-nya terisi."""
    for r in range(2, last_r + 1):
        if ws.cell(r, 1).value not in (None, ""):
            return r
    return 1

def sheet_heading(ws, grid_start):
    """Ambil 'Tabel X.X.X' + judul deskriptif dari area di atas grid.
    Pisah Indonesia/Inggris pada kemunculan tahun pertama (mis. '2025')."""
    num, chunks = "", []
    for r in range(1, grid_start):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if v is None: continue
            s = re.sub(r"\s+", " ", str(v).strip())
            if not s: continue
            if re.match(r"(?i)^tabel|^table", s) and len(s) < 22:
                num = num or s
            else:
                chunks.append(s)
    full = " ".join(chunks)
    m = re.search(r"20\d{2}(?:\s*[–-]\s*20\d{2})?", full)
    if m:
        indo = full[:m.end()].strip(" ,;")
        en = full[m.end():].strip(" ,;–-")
    else:
        indo, en = full, ""
    parts = ['<div class=xl-head>']
    if num:  parts.append(f'<div class=xl-no>{escape(num)}</div>')
    if indo: parts.append(f'<div class=xl-desc>{escape(indo)}</div>')
    if en:   parts.append(f'<div class=xl-en>{escape(en)}</div>')
    parts.append("</div>")
    return Markup("".join(parts))

HEADER_KW = ("DESA/KELURAHAN", "VILLAGE", "TINGKAT", "JENIS", "SUBDISTRICT")

def find_markers(ws, g0, last_r, last_c):
    """Semua baris penanda berisi '(1)' beserta peta kolom->nomor."""
    ms = []
    for r in range(g0, last_r + 1):
        nums = {}
        for c in range(1, last_c + 1):
            mm = re.fullmatch(r"\((\d+)\)", cellval(ws.cell(r, c).value).strip())
            if mm: nums[c] = int(mm.group(1))
        if nums and 1 in nums.values():
            ms.append((r, nums))
    return ms

def render_merged(ws, g0, last_r, last_c, markers, cval, cover2anchor):
    """Gabungkan blok bertumpuk (kode desa sama) menjadi satu tabel lebar,
    dengan header bertingkat (grup + daun) tetap dipertahankan."""
    blocks, prev_end = [], g0 - 1
    for i, (mrow, nums) in enumerate(markers):
        vcols = sorted(c for c, n in nums.items() if n >= 2)
        if not vcols: return None
        htop = g0 if i == 0 else prev_end + 1
        limit = markers[i+1][0] if i+1 < len(markers) else last_r + 1
        data, order, total = {}, [], None
        for rr in range(mrow + 1, limit):
            a = cellval(ws.cell(rr, 1).value).strip()
            if re.fullmatch(r"\d{3}", a):
                data[a] = rr; order.append(a)
            elif a:
                if any(k in a.upper() for k in HEADER_KW): break   # header blok berikutnya
                if order and total is None: total = rr; break       # baris total (mis. nama kec)
        prev_end = max([mrow] + list(data.values()) + ([total] if total else []))
        labels = {}
        for c in vcols:
            lab = ""
            for rr in range(mrow - 1, htop - 1, -1):
                v = cval(rr, c)
                if v is not None and str(v).strip(): lab = str(v); break
            labels[c] = lab
        blocks.append(dict(mrow=mrow, vcols=vcols, nums=nums, idc=min(vcols)-1,
                           data=data, order=order, total=total, labels=labels, htop=htop))
    b0 = blocks[0]
    if not b0["order"]: return None
    for b in blocks[1:]:                       # harus berbagi kode desa
        if not (set(b["order"]) & set(b0["order"])): return None
    idc = b0["idc"]
    idlabel = ""
    for rr in range(b0["mrow"] - 1, b0["htop"] - 1, -1):
        v = cval(rr, 1)
        if v is not None and str(v).strip(): idlabel = str(v); break

    def vcell(v):
        txt = cellval(v)
        if txt == "": txt = "–"                      # sel data kosong -> no data
        if isinstance(v, (int, float)) or txt in ("-", "–"):
            return f'<td class="xn">{escape(txt)}</td>'
        return f'<td class="xl-lab">{biling(txt)}</td>'

    def levels_of(b, c):
        """Tingkat header kolom c pada blok b (atas->bawah, daun terakhir)."""
        lv, prev = [], None
        for rr in range(b["htop"], b["mrow"]):
            key = cover2anchor.get((rr, c), (rr, c))
            v = cval(rr, c)
            if v is None or str(v).strip() == "" or key == prev: continue
            lv.append((key, str(v))); prev = key
        return lv or [((b["mrow"], c), b["labels"].get(c, ""))]

    outcols = [(b, c) for b in blocks for c in b["vcols"]]
    levels = [levels_of(b, c) for b, c in outcols]
    ncol = len(outcols)
    D = max((len(l) for l in levels), default=1)

    out = ["<table class=xl>"]
    covered = set()
    for R in range(D):
        cells = []
        if R == 0:
            cells.append(f'<td class="xh" colspan={idc} rowspan={D}>{biling(idlabel)}</td>')
        j = 0
        while j < ncol:
            if (R, j) in covered: j += 1; continue
            lv = levels[j]
            if R > len(lv) - 1: j += 1; continue
            key, val = lv[R]
            cs = 1                          # grup teratas digabung per teks; level bawah per anchor
            while j + cs < ncol and R <= len(levels[j+cs]) - 1 and (
                    levels[j+cs][R][1] == val if R == 0 else levels[j+cs][R][0] == key):
                cs += 1
            rs = (D - R) if R == len(lv) - 1 else 1     # daun mengisi ke bawah
            span = (f" colspan={cs}" if cs > 1 else "") + (f" rowspan={rs}" if rs > 1 else "")
            cells.append(f'<td class="xh"{span}>{biling(val)}</td>')
            for rr in range(R, R + rs):
                for cc in range(j, j + cs): covered.add((rr, cc))
            j += cs
        out.append("<tr>" + "".join(cells) + "</tr>")
    numrow = [f'<td class="xh" colspan={idc}>(1)</td>']
    for b, c in outcols:
        numrow.append(f'<td class="xh">({b["nums"][c]})</td>')
    out.append("<tr>" + "".join(numrow) + "</tr>")
    for code in b0["order"]:
        r0 = b0["data"][code]
        cells = [f'<td class="xl-lab">{escape(cellval(ws.cell(r0,1).value))}</td>']
        if idc > 1:
            cells.append(f'<td class="xl-lab" colspan={idc-1}>{biling(cellval(cval(r0,2)))}</td>')
        for b in blocks:
            rr = b["data"].get(code)
            for c in b["vcols"]:
                cells.append(vcell(ws.cell(rr, c).value if rr else None))
        out.append("<tr>" + "".join(cells) + "</tr>")
    if b0["total"]:
        cells = [f'<td class="xl-lab" colspan={idc}>{escape(cellval(ws.cell(b0["total"],1).value))}</td>']
        for b in blocks:
            rr = b["total"]
            for c in b["vcols"]:
                cells.append(vcell(ws.cell(rr, c).value if rr else None))
        out.append("<tr>" + "".join(cells) + "</tr>")
    out.append("</table>")
    return Markup("".join(out))

def render_sheet(ws):
    last_r, last_c = bounds(ws)
    if last_r == 0: return Markup("")
    anchor, covered, cover2anchor = {}, set(), {}
    for m in ws.merged_cells.ranges:
        if m.min_row > last_r or m.min_col > last_c: continue
        anchor[(m.min_row, m.min_col)] = (m.max_row-m.min_row+1, m.max_col-m.min_col+1)
        for r in range(m.min_row, m.max_row+1):
            for c in range(m.min_col, m.max_col+1):
                if (r, c) != (m.min_row, m.min_col):
                    covered.add((r, c)); cover2anchor[(r, c)] = (m.min_row, m.min_col)
    g0 = grid_start_row(ws, last_r)
    heading = sheet_heading(ws, g0)
    def cval(r, c):
        if (r, c) in cover2anchor:
            ar, ac = cover2anchor[(r, c)]; return ws.cell(ar, ac).value
        return ws.cell(r, c).value

    markers = find_markers(ws, g0, last_r, last_c)
    if len(markers) >= 2:
        merged = render_merged(ws, g0, last_r, last_c, markers, cval, cover2anchor)
        if merged is not None:
            return Markup(heading + merged)
    marker = markers[0][0] if markers else 0

    out = ['<table class=xl>']
    for r in range(g0, last_r+1):
        row_covered = any((r, c) in covered for c in range(1, last_c+1))
        # baris "efektif kosong": semua sel & anchor merge yg menutupinya tak berisi
        anchors = {cover2anchor.get((r, c), (r, c)) for c in range(1, last_c+1)}
        eff_content = any(ws.cell(ar, ac).value not in (None, "") for ar, ac in anchors)
        if r > marker and not eff_content:      # lewati baris placeholder kosong
            continue
        tds = []
        for c in range(1, last_c+1):
            if (r, c) in covered: continue
            raw = ws.cell(r, c).value
            txt = cellval(raw)
            rs, cs = anchor.get((r, c), (1, 1))
            cs = min(cs, last_c - c + 1)      # clamp agar tak melewati batas grid
            rs = min(rs, last_r - r + 1)
            span = (f" rowspan={rs}" if rs > 1 else "") + (f" colspan={cs}" if cs > 1 else "")
            if marker and r <= marker: cls, inner = "xh", biling(txt)
            elif isinstance(raw, (int, float)) or txt in ("-", "–"): cls, inner = "xn", escape(txt)
            else: cls, inner = "xl-lab", biling(txt)
            tds.append(f'<td class="{cls}"{span}>{inner}</td>')
        if tds or row_covered:
            out.append("<tr>" + "".join(tds) + "</tr>")
    out.append("</table>")
    return Markup(heading + Markup("".join(out)))

def render_bab(folder, fname):
    """Render SEMUA sheet di file (urutan template)."""
    path = os.path.join(HASIL, folder, fname)
    if not os.path.exists(path): return []
    wb = load_wb(path)
    return [render_sheet(wb[s]) for s in wb.sheetnames]

app = Flask(__name__)

CSS = """
:root{--bg:#f6f7f9;--card:#fff;--ink:#1a2332;--mut:#6b7688;--line:#e6e9ef;--acc:#0d6e6e;--acc2:#0a5252}
*{box-sizing:border-box}body{margin:0;font:15px/1.5 -apple-system,Segoe UI,Roboto,sans-serif;background:var(--bg);color:var(--ink)}
a{color:var(--acc);text-decoration:none}a:hover{text-decoration:underline}
header{background:linear-gradient(135deg,#0d6e6e,#0a5252);color:#fff;padding:26px 20px}
.wrap{max-width:1280px;margin:0 auto;padding:0 20px}
header h1{margin:0;font-size:22px}header p{margin:4px 0 0;opacity:.85;font-size:14px}
.crumb{padding:14px 0;color:var(--mut);font-size:13px}
.grid{display:grid;grid-template-columns:repeat(auto-fill,minmax(150px,1fr));gap:12px;margin:8px 0 22px}
.stat{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:14px}
.stat .n{font-size:24px;font-weight:700;color:var(--acc)}.stat .l{font-size:12px;color:var(--mut);margin-top:2px}
.kecs{display:grid;grid-template-columns:repeat(auto-fill,minmax(200px,1fr));gap:10px}
.kec{background:var(--card);border:1px solid var(--line);border-radius:12px;padding:14px 16px;display:flex;justify-content:space-between;align-items:center}
.kec:hover{border-color:var(--acc)}.kec b{font-size:15px}.kec span{color:var(--mut);font-size:12px}
.card{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:14px;margin:14px 0;overflow-x:auto}
.sec{font-size:14px;font-weight:700;letter-spacing:.03em;color:var(--acc2);margin:28px 0 6px;padding-top:8px;border-bottom:2px solid var(--acc);padding-bottom:6px}
.tag{display:inline-block;background:#e7f2f2;color:var(--acc2);border-radius:6px;padding:2px 8px;font-size:11px;font-weight:600;margin-left:4px}
/* layout kecamatan: sidebar kiri sticky + konten */
.kv{display:flex;gap:20px;align-items:flex-start}
.kv-side{position:sticky;top:64px;margin-top:52px;width:190px;flex:none;max-height:calc(100vh - 80px);overflow:auto;background:var(--card);border:1px solid var(--line);border-radius:12px;padding:8px}
.kv-side h4{margin:6px 8px 8px;font-size:11px;letter-spacing:.05em;text-transform:uppercase;color:var(--mut)}
.kv-link{display:block;padding:7px 10px;border-radius:8px;font-size:13px;color:var(--ink)}
.kv-link:hover{background:#eef4f4;text-decoration:none}
.kv-link.on{background:var(--acc);color:#fff;font-weight:600}
.kv-main{flex:1;min-width:0}
@media(max-width:860px){.kv-side{display:none}}
.tabs{position:sticky;top:0;background:var(--bg);padding:12px 0;border-bottom:1px solid var(--line);z-index:5;display:flex;gap:8px;flex-wrap:wrap}
.tab{background:#fff;border:1px solid var(--line);border-radius:20px;padding:6px 16px;font-size:13px;font-weight:600;color:var(--acc2);cursor:pointer}
.tab:hover{border-color:var(--acc)}
.tab.active{background:var(--acc);color:#fff;border-color:var(--acc)}
/* judul tabel (di luar grid) */
.xl-head{margin:2px 2px 12px}
.xl-no{font-weight:700;font-size:12px;letter-spacing:.03em;color:var(--acc);text-transform:uppercase}
.xl-desc{font-weight:700;font-size:15px;color:var(--ink);line-height:1.35;margin-top:2px}
.xl-en{font-weight:400;font-style:italic;font-size:12.5px;color:var(--mut);line-height:1.3;margin-top:1px}
/* tabel Excel setia template */
table.xl{border-collapse:collapse;font-size:13px;min-width:100%}
table.xl td{border:1px solid #c9d2d2;padding:5px 9px;vertical-align:middle;text-align:center}
table.xl td.xh{background:#eef4f4;color:var(--acc2);font-weight:600;font-size:12px;line-height:1.25}
table.xl td.xl-lab{text-align:left;white-space:nowrap}
table.xl td.xn{text-align:right;font-variant-numeric:tabular-nums}
table.xl .en{font-weight:400;font-style:italic;color:var(--mut)}
footer{color:var(--mut);font-size:12px;text-align:center;padding:30px 0}
/* admin bar & tombol */
.admin{display:flex;align-items:center;gap:12px;flex-wrap:wrap;margin-top:10px}
.admin-info{font-size:13px;opacity:.9}
.btn{display:inline-block;background:#fff;color:var(--acc2);border:1px solid var(--line);border-radius:8px;padding:7px 14px;font-size:13px;font-weight:600;cursor:pointer}
.btn:hover{text-decoration:none;border-color:var(--acc)}
.btn-warn{background:#fdecec;color:#b03030;border-color:#f3c9c9}
.up-card{background:var(--card);border:1px solid var(--line);border-radius:14px;padding:22px;max-width:560px;margin:24px auto}
.up-card h3{margin:0 0 6px}
.up-card input[type=file]{display:block;margin:6px 0 14px;font-size:14px}
.up-note{color:var(--mut);font-size:12.5px;line-height:1.5}
.err{background:#fdecec;border:1px solid #f3c9c9;color:#b03030;border-radius:10px;padding:12px 14px;margin:12px 0;white-space:pre-wrap;font-size:12.5px}
"""

BASE_HTML = """<!doctype html><html lang=id><head><meta charset=utf-8>
<meta name=viewport content="width=device-width,initial-scale=1">
<title>{{title}}</title><style>{{css}}</style></head><body>
<header><div class=wrap><h1>📊 PODES 2025 — {{kab}}</h1>
<p>Dashboard Potensi Desa · isian otomatis dari file .sav</p>
{{admin|safe}}
</div></header>
<div class=wrap>{{body|safe}}
<footer>Sumber: Pendataan Potensi Desa (PODES) 2025 · BPS · Tampilan mengikuti template tabel</footer>
</div></body></html>"""

UPLOAD = """
<div class=up-card>
  <h3>Muat Data PODES 2025 Desa</h3>
  <div class=up-note>Upload <b>satu file .sav</b> (data desa). Sistem akan otomatis:
  membuat template per kecamatan sesuai daftar desa di data, mengisi tabel dari .sav,
  lalu menampilkan dashboard. Data yang tak ada di .sav (mis. penduduk Bab 3) tampil "–".</div>
  {{err|safe}}
  <form method=post action="/upload" enctype="multipart/form-data" onsubmit="this.querySelector('button').disabled=true;this.querySelector('button').textContent='Memproses… (mohon tunggu)'">
    <p style="margin-top:16px"><b>File .sav</b><br><input type=file name=sav accept=".sav" required></p>
    <button class=btn type=submit>Proses & Muat</button>
  </form>
</div>"""

INDEX = """
<div class=crumb>Ringkasan Kabupaten</div>
<div class=grid>
  <div class=stat><div class=n>{{n_kec}}</div><div class=l>Kecamatan</div></div>
  <div class=stat><div class=n>{{n_desa}}</div><div class=l>Desa/Kelurahan</div></div>
  <div class=stat><div class=n>{{sd}}</div><div class=l>Sekolah Dasar</div></div>
  <div class=stat><div class=n>{{pusk}}</div><div class=l>Desa Ada Puskesmas</div></div>
  <div class=stat><div class=n>{{banjir}}</div><div class=l>Desa Alami Banjir</div></div>
  <div class=stat><div class=n>{{hotel}}</div><div class=l>Desa Ada Hotel</div></div>
  <div class=stat><div class=n>{{kop}}</div><div class=l>Total Koperasi</div></div>
  <div class=stat><div class=n>{{mini}}</div><div class=l>Minimarket</div></div>
</div>
<div class=crumb>Pilih Kecamatan</div>
<div class=kecs>
{% for k in kecs %}<a class=kec href="/kec/{{k.slug}}"><b>{{k.nama}}</b><span>{{k.n}} desa ›</span></a>{% endfor %}
</div>"""

KEC = """
<div class=kv>
<aside class=kv-side>
  <h4>Kecamatan</h4>
  {% for k in kecs %}<a class="kv-link{{' on' if k.on else ''}}" href="/kec/{{k.slug}}">{{k.nama}}</a>{% endfor %}
</aside>
<section class=kv-main>
<div class=crumb><a href="/">← Kabupaten</a> / <b>{{nama}}</b> · {{n_desa}} desa</div>
<div class=tabs>
  {% for s in sections %}<button class="tab{{' active' if loop.first else ''}}"
     onclick="showBab('{{s.id}}',this)">{{s.short}}</button>{% endfor %}
</div>
{% for s in sections %}
<div class="bab-panel" id="{{s.id}}"{{'' if loop.first else ' style=display:none'|safe}}>
  <div class=sec>{{s.title}}</div>
  {% if s.cards %}{% for c in s.cards %}<div class=card>{{c}}</div>{% endfor %}
  {% else %}<div class=card><em style="color:#6b7688">File belum tersedia.</em></div>{% endif %}
</div>
{% endfor %}
</section>
</div>
<script>
function showBab(id,btn){
  document.querySelectorAll('.bab-panel').forEach(function(p){p.style.display='none'});
  document.getElementById(id).style.display='';
  document.querySelectorAll('.tab').forEach(function(t){t.classList.remove('active')});
  btn.classList.add('active');
  window.scrollTo(0,0);
}
</script>
"""

def admin_bar():
    if STATE["df"] is None: return ""
    name = os.path.basename(STATE["sav"] or "")
    return (f'<div class=admin><span class=admin-info>📂 {escape(name)} · '
            f'{len(STATE["kecs"])} kecamatan · {len(STATE["df"])} desa</span>'
            f'<a class=btn href="/compile">📦 Compile ZIP</a>'
            f'<form method=post action="/clear" style="display:inline" '
            f'onsubmit="return confirm(\'Hapus project & hasil? Untuk pindah ke file PODES lain.\')">'
            f'<button class="btn btn-warn" type=submit>🗑 Clear Project</button></form></div>')

def page(title, body):
    return render_template_string(BASE_HTML, title=title, css=CSS, body=body,
                                  kab=STATE["kab"] or "PODES 2025", admin=admin_bar())

@app.route("/")
def index():
    if STATE["df"] is None:
        return page("Upload PODES", render_template_string(UPLOAD, err=""))
    df = STATE["df"]
    kecs = [{"slug": slug(k), "nama": k.title(), "n": len(sub_of(k))} for k in STATE["kecs"]]
    body = render_template_string(INDEX, kecs=kecs, n_kec=len(STATE["kecs"]), n_desa=len(df),
        sd=tot("r701ck2")+tot("r701ck3"),
        pusk=int(((df["r702dk2"].fillna(0)+df["r702ek2"].fillna(0))>0).sum()),
        banjir=int((df["r601bk2"]==1).sum()),
        hotel=int((df["r905hk2"].fillna(0)>0).sum()),
        kop=sum(tot(v) for v in ["r903a","r903b","r903c","r903d"]),
        mini=tot("r905ek2"))
    return page(f"PODES 2025 {STATE['kab']}", body)

@app.route("/upload", methods=["POST"])
def upload():
    f = request.files.get("sav")
    if not f or not f.filename.lower().endswith(".sav"):
        return page("Upload PODES", render_template_string(UPLOAD,
                    err='<div class=err>File .sav wajib diunggah.</div>'))
    savpath = os.path.join(WORK, "data.sav")
    f.save(savpath)
    try:
        process_sav(savpath)
    except Exception as e:
        return page("Upload PODES", render_template_string(UPLOAD,
                    err=f'<div class=err><b>Gagal memproses:</b>\n{escape(str(e))}</div>'))
    return redirect("/")

@app.route("/compile")
def compile_zip():
    if STATE["df"] is None or not os.path.isdir(HASIL): return redirect("/")
    mem = io.BytesIO()
    with zipfile.ZipFile(mem, "w", zipfile.ZIP_DEFLATED) as z:
        for root, _, files in os.walk(HASIL):
            for fn in files:
                fp = os.path.join(root, fn)
                z.write(fp, os.path.join("tabel terisi", os.path.relpath(fp, HASIL)))
    mem.seek(0)
    fn = f"PODES2025_{(STATE['kab'] or 'tabel').replace(' ','_')}_terisi.zip"
    return send_file(mem, download_name=fn, as_attachment=True, mimetype="application/zip")

@app.route("/clear", methods=["POST"])
def clear():
    clear_state()
    return redirect("/")

@app.route("/kec/<s>")
def kec(s):
    if STATE["df"] is None or s not in STATE["slug2kec"]: abort(404)
    kecname = STATE["slug2kec"][s]
    folder = STATE["kec2folder"].get(kecname)
    sections = []
    for bid, short, title, fname in BABS:
        cards = render_bab(folder, fname) if folder else []
        sections.append({"id": bid, "short": short, "title": title, "cards": cards})
    kecs = [{"slug": slug(k), "nama": k.title(), "on": k == kecname} for k in STATE["kecs"]]
    body = render_template_string(KEC, nama=kecname.title(), kecs=kecs,
                                  n_desa=len(sub_of(kecname)), sections=sections)
    return page(f"{kecname.title()} — PODES 2025", body)

@app.route("/healthz")
def health(): return "ok"

if __name__ == "__main__":
    app.run(host="127.0.0.1", port=int(os.environ.get("PORT", 8090)))
