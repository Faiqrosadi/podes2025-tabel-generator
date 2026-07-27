#!/usr/bin/env python3
"""Generate folder template per-kecamatan dari SATU .sav PODES 2025 Desa.

Struktur tabel BPS seragam nasional; yang beda hanya nama kab/kec & daftar desa.
Skrip ini: kloning master (per jumlah desa) -> sesuaikan jumlah baris desa ->
tulis kode+nama desa dari .sav -> ganti nama kabupaten/kecamatan di judul ->
HAPUS semua angka hardcode (data Bab 3, batas 1.1.3, rumus) sehingga bersih.

Hasil: folder template baru (default workspace/template_gen/) yang siap diisi
oleh pipeline fill_* (yang membaca data dari .sav).

Pakai:
  python3 generate.py <sav> [master_template_dir] [output_template_dir]
Default master = 'template tabel/', output = 'workspace/template_gen/'.
"""
import os, re, sys, glob, copy, shutil
import pyreadstat, openpyxl
from openpyxl.cell.cell import MergedCell

def norm(s): return re.sub(r"\s+", " ", str(s).strip()).upper()
def titlecase(s): return re.sub(r"\s+", " ", str(s).strip()).title()

def find_bands(ws):
    """Daftar (start,end) baris berturut-turut berkode desa 3 digit."""
    rows = [r for r in range(1, ws.max_row + 1)
            if re.fullmatch(r"\d{3}", str(ws.cell(r, 1).value).strip() if ws.cell(r, 1).value else "")]
    bands = []
    for r in rows:
        if bands and r == bands[-1][1] + 1:
            bands[-1][1] = r
        else:
            bands.append([r, r])
    return [tuple(b) for b in bands]

def row_merges(ws, r):
    """Merge horizontal yg mulai di baris r (offset kolom, lebar)."""
    out = []
    for m in ws.merged_cells.ranges:
        if m.min_row == r and m.max_row == r and m.max_col > m.min_col:
            out.append((m.min_col, m.max_col))
    return out

def copy_row(ws, src, dst):
    for c in range(1, ws.max_column + 1):
        s = ws.cell(src, c)
        d = ws.cell(dst, c)
        if not isinstance(d, MergedCell) and s.has_style:
            d._style = copy.copy(s._style)

def adjust_band(ws, band, target, name_merges):
    """Ubah jumlah baris band jadi `target` (delete/insert), jaga style+merge."""
    start, end = band
    cur = end - start + 1
    if cur == target:
        return
    if cur > target:
        ws.delete_rows(start + target, cur - target)
    else:
        add = target - cur
        ws.insert_rows(end + 1, add)
        for i in range(add):
            nr = end + 1 + i
            copy_row(ws, start, nr)
            for (c1, c2) in name_merges:
                try: ws.merge_cells(start_row=nr, start_column=c1, end_row=nr, end_column=c2)
                except Exception: pass

def marker_rows(ws):
    return [r for r in range(1, ws.max_row + 1)
            if any(str(ws.cell(r, c).value).strip() == "(1)" for c in range(1, ws.max_column + 1))]

def value_cols(ws, mrow):
    cols = []
    for c in range(1, ws.max_column + 1):
        m = re.fullmatch(r"\((\d+)\)", str(ws.cell(mrow, c).value).strip() if ws.cell(mrow, c).value else "")
        if m and int(m.group(1)) >= 2:
            cols.append(c)
    return cols

def strip_values(ws):
    """Hapus angka/rumus hardcode: sel nilai di bawah tiap penanda, + rumus, + 1.1.3."""
    # rumus di mana pun -> kosong
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and cell.value.strip().startswith("="):
                if not isinstance(cell, MergedCell): cell.value = None
    mrows = marker_rows(ws)
    for i, mrow in enumerate(mrows):
        vcols = value_cols(ws, mrow)
        end = mrows[i + 1] - 1 if i + 1 < len(mrows) else ws.max_row
        for r in range(mrow + 1, end + 1):
            for c in vcols:
                cell = ws.cell(r, c)
                if not isinstance(cell, MergedCell): cell.value = None
    # tabel batas/letak geografis (1.1.3) tanpa penanda: bersihkan nilai batas & teks bebas
    for r in range(1, ws.max_row + 1):
        a = str(ws.cell(r, 1).value or "").upper()
        if a.startswith(("SEBELAH UTARA", "SEBELAH TIMUR", "SEBELAH SELATAN", "SEBELAH BARAT")):
            for c in range(3, ws.max_column + 1):
                cell = ws.cell(r, c)
                if not isinstance(cell, MergedCell): cell.value = None
        txt = " ".join(str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1)).lower()
        if ("ketinggian" in txt and "permukaan laut" in txt) or ("jarak dari" in txt and "ibukota" in txt) \
           or ("terletak di sebelah" in txt):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if not isinstance(cell, MergedCell): cell.value = None

def rename_titles(ws, kab_from, kec_from, kab_to, kec_to):
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v.strip() and not isinstance(cell, MergedCell):
                nv = re.sub(kec_from, kec_to, v, flags=re.IGNORECASE)
                nv = re.sub(kab_from, kab_to, nv, flags=re.IGNORECASE)
                if nv != v: cell.value = nv

def gen_kecamatan(master_file, out_file, desa_list, kab_from, kec_from, kab_to, kec_to):
    wb = openpyxl.load_workbook(master_file)
    N = len(desa_list)
    for ws in wb.worksheets:
        bands = find_bands(ws)
        if bands:
            # deteksi merge nama desa dari baris desa pertama
            nmerges = row_merges(ws, bands[0][0])
            for band in reversed(bands):          # bawah->atas agar indeks stabil
                adjust_band(ws, band, N, nmerges)
            for band in find_bands(ws):           # tulis kode+nama desa
                for i in range(N):
                    r = band[0] + i
                    kode, nama = desa_list[i]
                    if not isinstance(ws.cell(r, 1), MergedCell): ws.cell(r, 1).value = kode
                    if not isinstance(ws.cell(r, 2), MergedCell): ws.cell(r, 2).value = nama
        strip_values(ws)
        rename_titles(ws, kab_from, kec_from, kab_to, kec_to)
    os.makedirs(os.path.dirname(out_file), exist_ok=True)
    wb.save(out_file)

def build_masters(master_dir):
    """count -> folder path, dari template master (berdasar jumlah baris desa Bab 2)."""
    m = {}
    for d in sorted(glob.glob(os.path.join(master_dir, "*"))):
        if not os.path.isdir(d): continue
        b2 = os.path.join(d, "Bab 2.xlsx")
        if not os.path.exists(b2): continue
        ws = openpyxl.load_workbook(b2)["Tabel 2.1.1"]
        n = sum(1 for b in find_bands(ws) for _ in range(b[0], b[1] + 1)) or \
            sum((b[1] - b[0] + 1) for b in find_bands(ws))
        bands = find_bands(ws)
        n = (bands[0][1] - bands[0][0] + 1) if bands else 0
        if n: m[n] = d
    return m

def pick_master(masters, n):
    """master dgn count == n; else terkecil yg >= n; else terbesar."""
    if n in masters: return masters[n], "exact"
    bigger = sorted(c for c in masters if c > n)
    if bigger: return masters[bigger[0]], f"delete dari {bigger[0]}"
    largest = max(masters)
    return masters[largest], f"insert dari {largest}"

def main():
    sav = sys.argv[1] if len(sys.argv) > 1 else None
    master_dir = sys.argv[2] if len(sys.argv) > 2 else os.path.join(os.path.dirname(__file__), "template tabel")
    out_dir = sys.argv[3] if len(sys.argv) > 3 else os.path.join(os.path.dirname(__file__), "workspace", "template_gen")
    if not sav or not os.path.exists(sav):
        sys.exit("Pakai: python3 generate.py <sav> [master_dir] [out_dir]")

    df, _ = pyreadstat.read_sav(sav)
    df["r104"] = df["r104"].astype(str).str.zfill(3)
    kab_to = titlecase(df["nama_kab"].iloc[0])
    masters = build_masters(master_dir)
    # nama kab/kec master (asumsi 1 kabupaten di master, ambil dari data master? -> pakai konstanta)
    kab_from = "Jepara"   # kabupaten master
    print(f"Master counts tersedia: {sorted(masters)}")
    print(f"Generate untuk: {kab_to}\n")

    if os.path.isdir(out_dir): shutil.rmtree(out_dir)
    kecs = df.sort_values("r103").drop_duplicates("r103")["nama_kec"].tolist()
    for idx, kec in enumerate(kecs, 1):
        sub = df[df["nama_kec"] == kec].sort_values("r104")
        desa_list = [(r.r104, titlecase(r.nama_desa)) for r in sub.itertuples()]
        N = len(desa_list)
        master_folder, how = pick_master(masters, N)
        kec_from = titlecase(re.sub(r"^\d+\s*", "", os.path.basename(master_folder)))
        kec_to = titlecase(kec)
        out_folder = os.path.join(out_dir, f"{idx:03d} {kec_to}")
        for bab in range(1, 8):
            src = os.path.join(master_folder, f"Bab {bab}.xlsx")
            if os.path.exists(src):
                gen_kecamatan(src, os.path.join(out_folder, f"Bab {bab}.xlsx"),
                              desa_list, kab_from, kec_from, kab_to, kec_to)
        print(f"  {idx:03d} {kec_to:16} {N:2} desa  (master {os.path.basename(master_folder)}, {how})")
    print(f"\nSelesai -> {out_dir}")

if __name__ == "__main__":
    main()
