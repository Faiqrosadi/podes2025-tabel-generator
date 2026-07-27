#!/usr/bin/env python3
"""
Isi otomatis Bab 7 (Ekonomi) template PODES 2025 Kab. Jepara dari data .sav.

Tabel yang diisi (lihat MAPPING.md):
  7.1  Banyaknya Koperasi per desa       -> r903a..r903d (per desa + total)
  7.3  Desa dgn keberadaan Bank          -> count(r902a1/a2/a3 > 0)
  7.4  Desa dgn keberadaan Koperasi      -> count(r903a/b/c/d > 0)
  7.5  Desa dgn keberadaan sarana dagang -> count(r905a/b/c/e/f k2 > 0)

Aturan: hanya kolom 2025; nilai 0 / nihil ditulis "-" (konvensi BPS).
Hasil ditulis ke folder hasil/<kecamatan>/Bab 7.xlsx (template asli tak diubah).
7.2 (sertifikat tanah) dilewati: data tidak ada di .sav.
"""
import os, re, glob, shutil
import pyreadstat
import openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
try:
    from config import SAV, TEMPLATE as TPL, OUTPUT as OUT
except Exception:
    SAV  = os.path.join(BASE, "data", "33_podes2025-desa_rev_20.sav")
    TPL  = os.path.join(BASE, "template tabel")
    OUT  = os.path.join(BASE, "hasil")

df, _ = pyreadstat.read_sav(SAV)

def norm(s):
    return re.sub(r"\s+", " ", str(s).strip()).upper()

def fmt(v):
    """0 / kosong -> '-', selain itu integer."""
    try:
        v = float(v)
    except (TypeError, ValueError):
        return "-"
    if v != v or v == 0:      # NaN atau 0
        return "-"
    return int(v) if v == int(v) else v

def count_desa(sub, var):
    """Jumlah desa dengan var > 0."""
    return int((sub[var].fillna(0) > 0).sum())

def find_rows(ws, label_col=1, start=1, end=None):
    """Kembalikan {baris: teks_kolom_A} untuk baris yang ada isinya."""
    end = end or ws.max_row
    out = {}
    for r in range(start, end + 1):
        v = ws.cell(r, label_col).value
        if v is not None and str(v).strip():
            out[r] = norm(v)
    return out

# ---- pengisi per tabel -------------------------------------------------------

def fill_71(ws, sub):
    """Per desa: kolom C,D,E,F = r903a,b,c,d ; baris Jumlah = total."""
    by_code = {str(row.r104): row for row in sub.itertuples()}
    varcols = {"C": "r903a", "D": "r903b", "E": "r903c", "F": "r903d"}
    totals = {c: 0.0 for c in varcols}
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if a is None:
            continue
        a = str(a).strip()
        if re.fullmatch(r"\d{3}", a):                 # baris desa
            row = by_code.get(a)
            for col, var in varcols.items():
                val = getattr(row, var, 0) if row is not None else 0
                val = 0 if (val is None or val != val) else float(val)
                totals[col] += val
                ws[f"{col}{r}"] = fmt(val)
        elif "JUMLAH" in norm(a) or "TOTAL" in norm(a):  # baris total
            for col in varcols:
                ws[f"{col}{r}"] = fmt(totals[col])

def fill_ringkas(ws, sub, rules, val_col="B"):
    """Tabel ringkasan: cari baris via keyword -> isi count desa di val_col."""
    rows = find_rows(ws)
    for r, text in rows.items():
        for var, keys, anti in rules:
            if any(k in text for k in keys) and not any(a in text for a in anti):
                ws[f"{val_col}{r}"] = fmt(count_desa(sub, var))
                break

RULES_73 = [   # bank
    ("r902a1", ["PEMERINTAH"], []),
    ("r902a2", ["SWASTA"], []),
    ("r902a3", ["PERKREDITAN", "BPR"], []),
]
RULES_74 = [   # koperasi
    ("r903a", ["UNIT DESA", "KUD"], []),
    ("r903b", ["INDUSTRI", "KERAJINAN"], []),
    ("r903c", ["SIMPAN PINJAM", "KOSPIN"], []),
    ("r903d", ["LAINNYA"], []),
]
RULES_75 = [   # sarana perdagangan
    ("r905ak2", ["PERTOKOAN"], []),
    ("r905ck2", ["SEMI PERMANEN"], []),
    ("r905bk2", ["PERMANEN"], ["SEMI"]),
    ("r905ek2", ["MINI MARKET", "MINIMARKET", "SWALAYAN", "SUPERMARKET"], []),
    ("r905fk2", ["RESTORAN", "RUMAH MAKAN"], []),
]

def sheet(wb, tag):
    for s in wb.sheetnames:
        if tag in s.replace(" ", ""):
            return wb[s]
    return None

# ---- main --------------------------------------------------------------------

def process(folder):
    kec = norm(re.sub(r"^\d+\s*", "", os.path.basename(folder)))
    sub = df[df["nama_kec"].apply(norm) == kec].sort_values("r104")
    if sub.empty:
        print(f"  ! LEWAT {folder}: nama_kec '{kec}' tak ada di data")
        return False
    src = os.path.join(folder, "Bab 7.xlsx")
    if not os.path.exists(src):
        print(f"  ! LEWAT {folder}: tak ada 'Bab 7.xlsx'")
        return False
    dst_dir = os.path.join(OUT, os.path.basename(folder))
    os.makedirs(dst_dir, exist_ok=True)
    dst = os.path.join(dst_dir, "Bab 7.xlsx")
    shutil.copy2(src, dst)

    wb = openpyxl.load_workbook(dst)
    if sheet(wb, "7.1"): fill_71(sheet(wb, "7.1"), sub)
    if sheet(wb, "7.3"): fill_ringkas(sheet(wb, "7.3"), sub, RULES_73)
    if sheet(wb, "7.4"): fill_ringkas(sheet(wb, "7.4"), sub, RULES_74)
    if sheet(wb, "7.5"): fill_ringkas(sheet(wb, "7.5"), sub, RULES_75)
    wb.save(dst)
    print(f"  OK {os.path.basename(folder):22} ({len(sub)} desa) -> {dst}")
    return True

if __name__ == "__main__":
    folders = sorted(glob.glob(os.path.join(TPL, "*")))
    folders = [f for f in folders if os.path.isdir(f)]
    print(f"Memproses {len(folders)} kecamatan...\n")
    ok = sum(process(f) for f in folders)
    print(f"\nSelesai: {ok}/{len(folders)} kecamatan. Output di: {OUT}")
