#!/usr/bin/env python3
"""Buat 'template master/' generik (lepas dari Jepara) dari 'template tabel/'.

- Ambil 1 folder per JUMLAH DESA unik (master by count).
- Hapus semua angka hardcode (data Bab 3, batas 1.1.3, rumus) via strip_values.
- Kosongkan nama desa (kolom B) -> diisi ulang saat generate.
- Ganti nama kecamatan di judul -> token {{KEC}}, sisa 'Jepara' -> {{KAB}}.

Hasil: template master/desa-<NN>/Bab 1..7.xlsx  (folder generik).
"""
import os, re, glob, shutil
import openpyxl
from openpyxl.cell.cell import MergedCell
from generate import strip_values, find_bands

SRC = os.path.join(os.path.dirname(__file__), "template tabel")
DST = os.path.join(os.path.dirname(__file__), "template master")

def desa_count(folder):
    ws = openpyxl.load_workbook(os.path.join(folder, "Bab 2.xlsx"))["Tabel 2.1.1"]
    bands = find_bands(ws)
    return (bands[0][1] - bands[0][0] + 1) if bands else 0

def kec_of(folder):
    return re.sub(r"^\d+\s*", "", os.path.basename(folder)).strip()

def clean_sheet(ws, kec_name):
    # 1) hapus angka/rumus/batas
    strip_values(ws)
    # 2) kosongkan nama desa (kolom B) di baris berkode desa
    for band in find_bands(ws):
        for r in range(band[0], band[1] + 1):
            cell = ws.cell(r, 2)
            if not isinstance(cell, MergedCell): cell.value = None
    # 3) tokenkan nama kecamatan -> {{KEC}}, sisa 'Jepara' -> {{KAB}}
    for row in ws.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and not isinstance(cell, MergedCell):
                v = re.sub(re.escape(kec_name), "{{KEC}}", cell.value, flags=re.IGNORECASE)
                v = re.sub(r"Jepara", "{{KAB}}", v, flags=re.IGNORECASE)
                if v != cell.value: cell.value = v

def main():
    if os.path.isdir(DST): shutil.rmtree(DST)
    # pilih 1 folder per jumlah desa
    by_count = {}
    for d in sorted(glob.glob(os.path.join(SRC, "*"))):
        if not os.path.isdir(d): continue
        n = desa_count(d)
        by_count.setdefault(n, d)          # ambil folder pertama utk tiap count
    print(f"Jumlah desa unik -> folder sumber: "
          f"{ {n: os.path.basename(f) for n,f in sorted(by_count.items())} }\n")
    for n, folder in sorted(by_count.items()):
        kec = kec_of(folder)
        out = os.path.join(DST, f"desa-{n:02d}")
        for bab in range(1, 8):
            src = os.path.join(folder, f"Bab {bab}.xlsx")
            if not os.path.exists(src): continue
            wb = openpyxl.load_workbook(src)
            for ws in wb.worksheets:
                clean_sheet(ws, kec)
            os.makedirs(out, exist_ok=True)
            wb.save(os.path.join(out, f"Bab {bab}.xlsx"))
        print(f"  desa-{n:02d}  (dari {os.path.basename(folder)}, kec '{kec}' -> token)")
    print(f"\nSelesai -> {DST}")

if __name__ == "__main__":
    main()
