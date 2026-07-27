#!/usr/bin/env python3
"""
Isi otomatis Bab 4 (Sosial) template PODES 2025 Kab. Jepara dari data .sav.
Lihat MAPPING.md. Hanya kolom 2025; nilai 0/nihil -> "-" (konvensi BPS).
Hasil -> hasil/<kecamatan>/Bab 4.xlsx (template asli tak diubah).

Tabel yang diisi:
  4.1.1  Desa punya fasilitas pendidikan (count desa, r701 negeri+swasta>0)
  4.1.2  Jumlah sekolah Negeri/Swasta/Total (Sum r701k2/k3) - hanya kolom 2025/2026
  4.2.1  Desa punya sarana kesehatan (count desa r702k2>0) - kolom 2025
  4.2.2  Jumlah sarana kesehatan per desa (r702/r703) - kolom yg tak ada di sav -> "-"
  4.3.1  Desa menurut sumber penerangan jalan (r502b) - kolom 2025
  4.3.2  Desa menurut bahan bakar memasak (r503)
  4.4.1  Desa mengalami bencana (r601k2==1) - kolom 2025
  4.4.3  Desa punya upaya mitigasi bencana (r602==1) - kolom 2025
DILEWATI: 4.1.3-4.1.35 (murid/guru/ruang kelas), 4.2.3-4.2.6, 4.4.2 (TBD), 4.5.x
"""
import os, re, glob, shutil
import pyreadstat
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

def safe_set(ws, coord, value):
    if not isinstance(ws[coord], MergedCell):
        ws[coord] = value

BASE = os.path.dirname(os.path.abspath(__file__))
try:
    from config import SAV, TEMPLATE as TPL, OUTPUT as OUT
except Exception:
    SAV  = os.path.join(BASE, "data", "33_podes2025-desa_rev_20.sav")
    TPL  = os.path.join(BASE, "template tabel")
    OUT  = os.path.join(BASE, "hasil")

df, _ = pyreadstat.read_sav(SAV)
df["r104"] = df["r104"].astype(str)

def norm(s): return re.sub(r"\s+", " ", str(s).strip()).upper()
def fmt(v):
    try: v = float(v)
    except (TypeError, ValueError): return "-"
    if v != v or v == 0: return "-"
    return int(v) if v == int(v) else v
def cnt(sub, cond): return int(cond.sum())
def num(sub, var): return sub[var].fillna(0)

def sheet(wb, tag):
    for s in wb.sheetnames:
        if tag in s.replace(" ", ""):
            return wb[s]
    return None

def row_labels(ws, col=1):
    out = {}
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, col).value
        if v is not None and str(v).strip() and not str(v).strip().startswith("("):
            out[r] = norm(v)
    return out

# ---------- jenjang pendidikan (kode -> huruf r701) ----------
JENJANG = [   # (token di label, huruf r701)
    ("(TK)", "a"), ("(RA)", "b"), ("(SD)", "c"), ("(MI)", "d"),
    ("(SMP)", "e"), ("(MTS)", "f"), ("(SMA)", "g"), ("(SMK)", "i"),
    ("(MA)", "h"), ("AKADEMI", "j"),
]
def jenjang_of(text):
    for tok, h in JENJANG:
        if tok in text:
            return h
    return None

# ---------- Tabel 4.1.1 : count desa punya jenjang (col G) ----------
def fill_411(ws, sub):
    for r, t in row_labels(ws).items():
        h = jenjang_of(t)
        if h:
            neg, swa = num(sub, f"r701{h}k2"), num(sub, f"r701{h}k3")
            ws[f"G{r}"] = fmt(cnt(sub, (neg + swa) > 0))

# ---------- Tabel 4.1.2 : jumlah sekolah (E=Negeri, G=Swasta, I=Total 2025/26) ----------
def fill_412(ws, sub):
    for r, t in row_labels(ws).items():
        h = jenjang_of(t)
        if h:
            neg = int(num(sub, f"r701{h}k2").sum())
            swa = int(num(sub, f"r701{h}k3").sum())
            ws[f"E{r}"], ws[f"G{r}"], ws[f"I{r}"] = fmt(neg), fmt(swa), fmt(neg + swa)
            ws[f"D{r}"] = ws[f"F{r}"] = ws[f"H{r}"] = "-"   # tahun 2024/2025 tak tersedia

# ---------- Tabel 4.2.1 : count desa punya faskes (col G) ----------
FASKES_421 = [
    (["RUMAH SAKIT"], [], "a"),
    (["TANPA RAWAT INAP"], [], "e"),
    (["RAWAT INAP"], ["TANPA"], "d"),
    (["APOTEK"], [], "l"),
]
def fill_421(ws, sub):
    for r, t in row_labels(ws).items():
        for keys, anti, h in FASKES_421:
            if any(k in t for k in keys) and not any(a in t for a in anti):
                ws[f"G{r}"] = fmt(cnt(sub, num(sub, f"r702{h}k2") > 0)); break

# ---------- Tabel 4.2.2 : jumlah faskes per desa (3 blok, marker (2)..(14)) ----------
# nomor kolom (n) -> variabel; None = tak ada di .sav
COL_NUM_VAR = {2:"r702ak2",3:"r702dk2",4:"r702ek2",5:"r702fk2",6:None,7:"r703a",
               8:None,9:"r702hk2",10:None,11:None,12:None,13:"r702lk2",14:"r702mk2"}
def fill_422(ws, sub):
    kec = norm(sub.iloc[0]["nama_kec"])
    by_code = {str(r.r104): r for r in sub.itertuples()}
    # temukan marker rows (baris berisi "(1)") -> peta kolom->nomor
    for r in range(1, ws.max_row + 1):
        rowvals = {c: ws.cell(r, c).value for c in range(1, ws.max_column + 1)}
        if any(str(v).strip() == "(1)" for v in rowvals.values()):
            colmap = {}   # col_letter -> nomor
            for c, v in rowvals.items():
                m = re.fullmatch(r"\((\d+)\)", str(v).strip()) if v is not None else None
                if m and int(m.group(1)) >= 2:
                    colmap[c] = int(m.group(1))
            # isi baris data di bawah marker sampai baris total (nama kec)
            totals = {c: 0.0 for c in colmap}
            rr = r + 1
            while rr <= ws.max_row:
                a = ws.cell(rr, 1).value
                a = str(a).strip() if a is not None else ""
                if re.fullmatch(r"\d{3}", a):
                    row = by_code.get(a)
                    for c, n in colmap.items():
                        var = COL_NUM_VAR.get(n)
                        cl = openpyxl.utils.get_column_letter(c)
                        if var is None:
                            ws[f"{cl}{rr}"] = "-"
                        else:
                            val = getattr(row, var, 0) if row is not None else 0
                            val = 0 if (val is None or val != val) else float(val)
                            totals[c] += val
                            ws[f"{cl}{rr}"] = fmt(val)
                elif norm(a) == kec or "JUMLAH" in norm(a) or "TOTAL" in norm(a):
                    for c, n in colmap.items():
                        cl = openpyxl.utils.get_column_letter(c)
                        ws[f"{cl}{rr}"] = "-" if COL_NUM_VAR.get(n) is None else fmt(totals[c])
                    break
                elif a and not re.fullmatch(r"\d{3}", a):
                    break   # masuk blok/hdr berikutnya
                rr += 1

# ---------- Tabel 4.3.1 : sumber penerangan jalan (col G, r502b) ----------
def fill_431(ws, sub):
    for r, t in row_labels(ws).items():
        code = None
        if "NON-PEMERINTAH" in t or "NON PEMERINTAH" in t: code = 2
        elif "NON LISTRIK" in t or "NON-LISTRIK" in t: code = 3
        elif "PEMERINTAH" in t: code = 1
        if code:
            ws[f"G{r}"] = fmt(cnt(sub, num(sub, "r502b") == code))

# ---------- Tabel 4.3.2 : bahan bakar memasak (col F, r503) ----------
BBM = [(["5,5 KG"],2),(["12 KG"],3),(["3 KG"],4),(["GAS KOTA","CITY GAS"],5),
       (["BIOGAS"],6),(["MINYAK TANAH","KEROSENE"],7),(["BRIKET"],8),
       (["ARANG","CHOCOAL"],9),(["KAYU","FIREWOOD"],10),(["LAINNYA","OTHERS"],11),
       (["LISTRIK","ELECTRIC"],1)]
def fill_432(ws, sub):
    for r, t in row_labels(ws).items():
        for keys, code in BBM:
            if any(k in t for k in keys):
                ws[f"F{r}"] = fmt(cnt(sub, num(sub, "r503") == code)); break

# ---------- Tabel 4.4.1 : desa mengalami bencana (col F, r601k2==1) ----------
BENCANA = [(["GEMPA"],"d"),(["TSUNAMI"],"e"),(["GUNUNG","MELETUS"],"h"),
           (["LONGSOR"],"a"),(["BANJIR BANDANG"],"c"),
           (["KEKERINGAN","DROUGHT"],"j"),(["KEBAKARAN","KARHUTLA"],"i"),
           (["ANGIN","PUYUH","PUTING"],"g"),(["GELOMBANG","TIDAL"],"f"),
           (["ABRASI"],"k"),(["BANJIR"],"b")]  # BANJIR terakhir (agar Bandang tertangkap dulu)
def fill_441(ws, sub):
    for r, t in row_labels(ws).items():
        for keys, h in BENCANA:
            if any(k in t for k in keys):
                ws[f"F{r}"] = fmt(cnt(sub, num(sub, f"r601{h}k2") == 1)); break

# ---------- Tabel 4.4.2 : desa terdapat korban jiwa (F=2024 k4, G=2025 k7) ----------
def fill_442(ws, sub):
    for r, t in row_labels(ws).items():
        for keys, h in BENCANA:
            if any(k in t for k in keys):
                ws[f"F{r}"] = fmt(cnt(sub, num(sub, f"r601{h}k4") > 0))   # meninggal 2024
                ws[f"G{r}"] = fmt(cnt(sub, num(sub, f"r601{h}k7") > 0))   # meninggal 2025
                break

# ---------- Tabel 4.4.3 : mitigasi bencana (col G, r602==1) ----------
def fill_443(ws, sub):
    for r, t in row_labels(ws).items():
        v = None
        if "PERINGATAN DINI" in t and "TSUNAMI" in t: v = "r602b"
        elif "PERINGATAN DINI" in t: v = "r602a"
        elif "KESELAMATAN" in t or "PERLENGKAPAN" in t: v = "r602c"
        elif "RAMBU" in t or "EVAKUASI" in t: v = "r602d"
        elif "NORMALISASI" in t or "PERAWATAN" in t: v = "r602e"
        if v:
            ws[f"G{r}"] = fmt(cnt(sub, num(sub, v) == 1))

FILLERS = {"4.1.1":fill_411,"4.1.2":fill_412,"4.2.1":fill_421,"4.2.2":fill_422,
           "4.3.1":fill_431,"4.3.2":fill_432,"4.4.1":fill_441,"4.4.2":fill_442,"4.4.3":fill_443}

# ---------- Kolom "Sekolah" di 4.1.5-4.1.35 (per desa) dari r701 ----------
def jenjang_letter(u):
    if "RAUDHATUL" in u or "RAUDATUL" in u or "(RA)" in u: return "b"
    if "IBTIDAIYAH" in u or "(MI)" in u: return "d"
    if "TSANAWIYAH" in u or "(MTS)" in u: return "f"
    if "ALIYAH" in u or "(MA)" in u: return "h"
    if "TAMAN KANAK" in u or "(TK)" in u: return "a"
    if "SEKOLAH DASAR" in u or "(SD)" in u: return "c"
    if "SLTP" in u or "MENENGAH PERTAMA" in u or "(SMP)" in u: return "e"
    if "KEJURUAN" in u or "SMK" in u: return "i"
    if "SLTA" in u or "MENENGAH ATAS" in u or "(SMA)" in u: return "g"
    return None

def sheet_title(ws):
    t = ""
    for r in range(1, 4):
        for c in range(1, 6):
            v = ws.cell(r, c).value
            if v and isinstance(v, str) and len(v) > len(t): t = v
    return t

def fill_sekolah(ws, sub):
    """Isi kolom 'Sekolah' per desa (jumlah sekolah) dari r701<jenjang>k<2/3>."""
    up = norm(sheet_title(ws))
    jen = jenjang_letter(up)
    if not jen: return
    c2a = {}
    for m in ws.merged_cells.ranges:
        for r in range(m.min_row, m.max_row + 1):
            for c in range(m.min_col, m.max_col + 1):
                if (r, c) != (m.min_row, m.min_col): c2a[(r, c)] = (m.min_row, m.min_col)
    def cval(r, c):
        a = c2a.get((r, c), (r, c)); return ws.cell(*a).value
    marker = None
    for r in range(1, ws.max_row + 1):
        if any(str(ws.cell(r, c).value).strip() == "(1)" for c in range(1, ws.max_column + 1)):
            marker = r; break
    if not marker: return
    # status dari judul (bila tabel utuh negeri/swasta)
    tstatus = "negeri" if ("NEGERI" in up and "SWASTA" not in up) else \
              ("swasta" if ("SWASTA" in up and "NEGERI" not in up) else None)
    if tstatus is None and jen == "b": tstatus = "swasta"   # RA = swasta
    # kolom yg header daunnya "Sekolah"
    sek_cols = []
    for c in range(2, ws.max_column + 1):
        for r in range(1, marker):
            v = cval(r, c)
            if v and str(v).split("\n")[0].strip().lower() == "sekolah":
                sek_cols.append(c); break
    kec = norm(sub.iloc[0].nama_kec)
    by_code = {str(r.r104): r for r in sub.itertuples()}
    for c in sek_cols:
        st = tstatus
        if st is None:                       # tabel campur: cari Negeri/Swasta di header atas kolom
            for r in range(1, marker):
                vu = str(cval(r, c) or "").upper()
                if "NEGERI" in vu: st = "negeri"; break
                if "SWASTA" in vu: st = "swasta"; break
        if st is None: continue
        var = f"r701{jen}k{'2' if st == 'negeri' else '3'}"
        colL = get_column_letter(c); total = 0.0
        for r in range(marker + 1, ws.max_row + 1):
            a = ws.cell(r, 1).value
            if a is None: continue
            a = str(a).strip()
            if re.fullmatch(r"\d{3}", a):
                row = by_code.get(a)
                val = getattr(row, var, 0) if row is not None else 0
                val = 0 if (val is None or val != val) else float(val)
                total += val
                safe_set(ws, f"{colL}{r}", fmt(val))
            elif "JUMLAH" in a.upper() or "TOTAL" in a.upper() or norm(a) == kec:
                safe_set(ws, f"{colL}{r}", fmt(total))

def process(folder):
    kec = norm(re.sub(r"^\d+\s*", "", os.path.basename(folder)))
    sub = df[df["nama_kec"].apply(norm) == kec].sort_values("r104")
    if sub.empty:
        print(f"  ! LEWAT {folder}: nama_kec '{kec}' tak ada"); return False
    src = os.path.join(folder, "Bab 4.xlsx")
    if not os.path.exists(src):
        print(f"  ! LEWAT {folder}: tak ada 'Bab 4.xlsx'"); return False
    dst_dir = os.path.join(OUT, os.path.basename(folder)); os.makedirs(dst_dir, exist_ok=True)
    dst = os.path.join(dst_dir, "Bab 4.xlsx"); shutil.copy2(src, dst)
    wb = openpyxl.load_workbook(dst)
    done = []
    for tag, fn in FILLERS.items():
        ws = sheet(wb, tag)
        if ws is not None: fn(ws, sub); done.append(tag)
    # kolom "Sekolah" pada 4.1.5-4.1.35
    for s in wb.sheetnames:
        m = re.search(r"4\.1\.(\d+)", s.replace(" ", ""))
        if m and int(m.group(1)) >= 5:
            fill_sekolah(wb[s], sub)
    wb.save(dst)
    print(f"  OK {os.path.basename(folder):22} ({len(sub)} desa) [{','.join(done)}]")
    return True

if __name__ == "__main__":
    folders = [f for f in sorted(glob.glob(os.path.join(TPL, "*"))) if os.path.isdir(f)]
    print(f"Memproses {len(folders)} kecamatan...\n")
    ok = sum(process(f) for f in folders)
    print(f"\nSelesai: {ok}/{len(folders)} kecamatan. Output di: {OUT}")
